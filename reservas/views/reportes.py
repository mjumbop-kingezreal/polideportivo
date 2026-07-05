"""Vistas de reportes: dashboard y exportaciones administrativas."""

import io
import json
from datetime import timedelta

from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from ..models import Cancha, Reserva, Usuario
from .helpers import rol_requerido


def _safe_ratio(part, total):
    if not total:
        return 0
    return round((part / total) * 100, 1)


def _truncate(value, max_length=34):
    text = str(value or '-')
    if len(text) <= max_length:
        return text
    return f'{text[:max_length - 3]}...'


def _obtener_datos_reportes():
    generated_at = timezone.localtime()
    today = timezone.localdate()
    thirty_days_ago = today - timedelta(days=29)
    sport_labels = dict(Cancha.TipoDeporte.choices)

    booking_summary = Reserva.objects.aggregate(
        total=Count('id'),
        confirmed=Count('id', filter=Q(estado=Reserva.Estado.CONFIRMADA)),
        attended=Count('id', filter=Q(estado=Reserva.Estado.ASISTIDA)),
        cancelled=Count('id', filter=Q(estado=Reserva.Estado.CANCELADA)),
        missed=Count('id', filter=Q(estado=Reserva.Estado.NO_ASISTIDA)),
    )

    total_reservas = booking_summary['total'] or 0
    reservas_confirmadas = booking_summary['confirmed'] or 0
    reservas_asistidas = booking_summary['attended'] or 0
    reservas_canceladas = booking_summary['cancelled'] or 0
    reservas_no_asistidas = booking_summary['missed'] or 0

    total_usuarios = Usuario.objects.count()
    total_canchas = Cancha.objects.count()

    reservas_qs = Reserva.objects.select_related('usuario', 'cancha').order_by('-fecha_creacion')
    usuarios_qs = Usuario.objects.order_by('-fecha_creacion')
    canchas_qs = Cancha.objects.order_by('nombre')
    ultimas_reservas = list(reservas_qs[:20])

    reservas_por_dia = list(
        Reserva.objects.filter(fecha__gte=thirty_days_ago)
        .values('fecha')
        .annotate(total=Count('id'))
        .order_by('fecha')
    )
    chart_dias_labels = [item['fecha'].strftime('%d/%m') for item in reservas_por_dia]
    chart_dias_data = [item['total'] for item in reservas_por_dia]

    canchas_populares_raw = list(
        Reserva.objects.values('cancha__nombre')
        .annotate(total=Count('id'))
        .order_by('-total', 'cancha__nombre')[:5]
    )
    chart_canchas_labels = [item['cancha__nombre'] for item in canchas_populares_raw]
    chart_canchas_data = [item['total'] for item in canchas_populares_raw]
    top_canchas = [
        {
            'nombre': item['cancha__nombre'],
            'total': item['total'],
            'porcentaje': _safe_ratio(item['total'], total_reservas),
        }
        for item in canchas_populares_raw
    ]

    distribucion_estados_raw = list(
        Reserva.objects.values('estado')
        .annotate(total=Count('id'))
        .order_by('-total', 'estado')
    )
    estado_labels = dict(Reserva.Estado.choices)
    estado_breakdown = [
        {
            'estado': item['estado'],
            'label': estado_labels.get(item['estado'], item['estado']),
            'total': item['total'],
            'porcentaje': _safe_ratio(item['total'], total_reservas),
        }
        for item in distribucion_estados_raw
    ]

    distribucion_deportes_raw = list(
        Reserva.objects.values('cancha__tipo_deporte')
        .annotate(total=Count('id'))
        .order_by('-total', 'cancha__tipo_deporte')
    )
    chart_deporte_labels = [
        sport_labels.get(item['cancha__tipo_deporte'], item['cancha__tipo_deporte'])
        for item in distribucion_deportes_raw
    ]
    chart_deporte_data = [item['total'] for item in distribucion_deportes_raw]
    deporte_breakdown = [
        {
            'label': sport_labels.get(item['cancha__tipo_deporte'], item['cancha__tipo_deporte']),
            'total': item['total'],
            'porcentaje': _safe_ratio(item['total'], total_reservas),
        }
        for item in distribucion_deportes_raw
    ]

    reservas_ultimos_30 = sum(chart_dias_data)
    tasa_asistencia = _safe_ratio(reservas_asistidas, total_reservas)
    tasa_cancelacion = _safe_ratio(reservas_canceladas, total_reservas)
    promedio_diario = round(reservas_ultimos_30 / 30, 1) if reservas_ultimos_30 else 0
    cancha_destacada = top_canchas[0]['nombre'] if top_canchas else 'Sin datos'
    deporte_destacado = deporte_breakdown[0]['label'] if deporte_breakdown else 'Sin datos'

    resumen_metricas = [
        ('Total reservas', total_reservas),
        ('Reservas confirmadas', reservas_confirmadas),
        ('Reservas asistidas', reservas_asistidas),
        ('Reservas canceladas', reservas_canceladas),
        ('Reservas no asistidas', reservas_no_asistidas),
        ('Usuarios registrados', total_usuarios),
        ('Canchas registradas', total_canchas),
        ('Reservas ultimos 30 dias', reservas_ultimos_30),
        ('Promedio diario', promedio_diario),
    ]

    observaciones = [
        f'Tasa de asistencia actual: {tasa_asistencia}%.',
        f'Tasa de cancelacion actual: {tasa_cancelacion}%.',
        f'La cancha con mayor demanda es {cancha_destacada}.',
        f'El deporte con mayor movimiento es {deporte_destacado}.',
    ]

    return {
        'generated_at': generated_at,
        'total_reservas': total_reservas,
        'reservas_confirmadas': reservas_confirmadas,
        'reservas_canceladas': reservas_canceladas,
        'reservas_no_asistidas': reservas_no_asistidas,
        'reservas_asistidas': reservas_asistidas,
        'total_usuarios': total_usuarios,
        'total_canchas': total_canchas,
        'reservas_ultimos_30': reservas_ultimos_30,
        'tasa_asistencia': tasa_asistencia,
        'tasa_cancelacion': tasa_cancelacion,
        'promedio_diario': promedio_diario,
        'cancha_destacada': cancha_destacada,
        'deporte_destacado': deporte_destacado,
        'ultimas_reservas': ultimas_reservas,
        'top_canchas': top_canchas,
        'estado_breakdown': estado_breakdown,
        'deporte_breakdown': deporte_breakdown,
        'resumen_metricas': resumen_metricas,
        'observaciones': observaciones,
        'chart_dias_labels': chart_dias_labels,
        'chart_dias_data': chart_dias_data,
        'chart_canchas_labels': chart_canchas_labels,
        'chart_canchas_data': chart_canchas_data,
        'chart_deporte_labels': chart_deporte_labels,
        'chart_deporte_data': chart_deporte_data,
        'reservas_exportacion': reservas_qs,
        'usuarios_exportacion': usuarios_qs,
        'canchas_exportacion': canchas_qs,
    }


@rol_requerido('administrador', 'municipio')
def reportes_view(request):
    data = _obtener_datos_reportes()
    context = {
        **data,
        'usuario': request.usuario,
        'chart_dias_labels': json.dumps(data['chart_dias_labels']),
        'chart_dias_data': json.dumps(data['chart_dias_data']),
        'chart_canchas_labels': json.dumps(data['chart_canchas_labels']),
        'chart_canchas_data': json.dumps(data['chart_canchas_data']),
        'chart_deporte_labels': json.dumps(data['chart_deporte_labels']),
        'chart_deporte_data': json.dumps(data['chart_deporte_data']),
    }
    return render(request, 'reservas/reportes.html', context)


@rol_requerido('administrador', 'municipio')
def exportar_pdf_view(request):
    """Exporta un reporte PDF ejecutivo del dashboard."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = _obtener_datos_reportes()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.25 * cm,
        rightMargin=1.25 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.3 * cm,
    )

    brand = colors.HexColor('#F97316')
    brand_dark = colors.HexColor('#183B56')
    soft = colors.HexColor('#F8FAFC')
    border = colors.HexColor('#E2E8F0')
    muted = colors.HexColor('#64748B')
    success = colors.HexColor('#16A34A')
    danger = colors.HexColor('#EF4444')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'], fontName='Helvetica-Bold',
        fontSize=22, leading=26, textColor=brand_dark, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle', parent=styles['Normal'], fontSize=10,
        leading=14, textColor=muted, spaceAfter=12,
    )
    section_style = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'], fontName='Helvetica-Bold',
        fontSize=12, textColor=brand_dark, spaceAfter=8, spaceBefore=8,
    )
    note_style = ParagraphStyle(
        'NoteStyle', parent=styles['Normal'], fontSize=9.2, leading=12,
        textColor=brand_dark,
    )
    card_value_style = ParagraphStyle(
        'CardValue', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=16, leading=19, textColor=brand_dark, alignment=1,
    )
    card_label_style = ParagraphStyle(
        'CardLabel', parent=styles['Normal'], fontSize=8.2, leading=10.5,
        textColor=muted, alignment=1,
    )

    def build_table(rows, widths, align='LEFT'):
        table = Table(rows, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (0, 0), (-1, -1), align),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, soft]),
            ('GRID', (0, 0), (-1, -1), 0.6, border),
        ]))
        return table

    def draw_page_chrome(canvas, pdf_doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(brand_dark)
        canvas.drawString(pdf_doc.leftMargin, pdf_doc.pagesize[1] - 0.95 * cm, 'PoliReserva | Reporte ejecutivo de reservas')
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(muted)
        canvas.drawRightString(pdf_doc.pagesize[0] - pdf_doc.rightMargin, 0.8 * cm, f'Pagina {canvas.getPageNumber()}')
        canvas.restoreState()

    elements = [
        Paragraph('Reporte Ejecutivo del Polideportivo', title_style),
        Paragraph(
            f'Generado el {data["generated_at"].strftime("%d/%m/%Y %H:%M")}. Reporte orientado a analisis de demanda, comportamiento y seguimiento de reservas.',
            subtitle_style,
        ),
    ]

    cards = Table([
        [
            Paragraph(str(data['total_reservas']), card_value_style),
            Paragraph(f"{data['tasa_asistencia']}%", card_value_style),
            Paragraph(f"{data['tasa_cancelacion']}%", card_value_style),
            Paragraph(str(data['reservas_ultimos_30']), card_value_style),
        ],
        [
            Paragraph('Total reservas', card_label_style),
            Paragraph('Tasa de asistencia', card_label_style),
            Paragraph('Tasa de cancelacion', card_label_style),
            Paragraph('Reservas ultimos 30 dias', card_label_style),
        ]
    ], colWidths=[6.6 * cm, 6.6 * cm, 6.6 * cm, 6.6 * cm])
    cards.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.8, border),
        ('INNERGRID', (0, 0), (-1, -1), 0.6, border),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.extend([cards, Spacer(1, 0.45 * cm)])

    elements.append(Paragraph('Observaciones clave', section_style))
    observaciones_rows = [[Paragraph(f'• {item}', note_style)] for item in data['observaciones']]
    observaciones_table = Table(observaciones_rows, colWidths=[26.4 * cm])
    observaciones_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.8, border),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, soft]),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.extend([observaciones_table, Spacer(1, 0.35 * cm)])

    elements.append(Paragraph('Resumen general', section_style))
    resumen_rows = [['Metrica', 'Valor']]
    for label, value in data['resumen_metricas']:
        resumen_rows.append([label, str(value)])
    elements.extend([build_table(resumen_rows, [15 * cm, 3.8 * cm], align='CENTER'), Spacer(1, 0.35 * cm)])

    elements.append(Paragraph('Estados de reserva', section_style))
    estado_rows = [['Estado', 'Reservas', 'Participacion']]
    for item in data['estado_breakdown']:
        estado_rows.append([item['label'], str(item['total']), f"{item['porcentaje']}%"])
    if len(estado_rows) == 1:
        estado_rows.append(['Sin datos', '0', '0%'])
    elements.extend([build_table(estado_rows, [10 * cm, 4 * cm, 4.8 * cm], align='CENTER'), Spacer(1, 0.35 * cm)])

    elements.append(Paragraph('Deportes con mayor movimiento', section_style))
    deporte_rows = [['Deporte', 'Reservas', 'Participacion']]
    for item in data['deporte_breakdown']:
        deporte_rows.append([item['label'], str(item['total']), f"{item['porcentaje']}%"])
    if len(deporte_rows) == 1:
        deporte_rows.append(['Sin datos', '0', '0%'])
    elements.extend([build_table(deporte_rows, [10 * cm, 4 * cm, 4.8 * cm], align='CENTER'), Spacer(1, 0.35 * cm)])

    elements.append(Paragraph('Canchas mas reservadas', section_style))
    canchas_rows = [['Cancha', 'Reservas', 'Participacion']]
    for cancha in data['top_canchas']:
        canchas_rows.append([_truncate(cancha['nombre']), str(cancha['total']), f"{cancha['porcentaje']}%"])
    if len(canchas_rows) == 1:
        canchas_rows.append(['Sin datos', '0', '0%'])
    elements.extend([build_table(canchas_rows, [12 * cm, 3.4 * cm, 3.4 * cm], align='CENTER'), Spacer(1, 0.35 * cm)])

    elements.append(Paragraph('Ultimas reservas registradas', section_style))
    reservas_rows = [['ID', 'Usuario', 'Cancha', 'Fecha', 'Horario', 'Estado']]
    for reserva in data['ultimas_reservas']:
        estado_color = success if reserva.estado in (Reserva.Estado.CONFIRMADA, Reserva.Estado.ASISTIDA) else danger
        reservas_rows.append([
            f'#{reserva.id}',
            _truncate(reserva.usuario.nombre, 22),
            _truncate(reserva.cancha.nombre, 22),
            reserva.fecha.strftime('%d/%m/%Y'),
            f'{reserva.hora_inicio.strftime("%H:%M")} - {reserva.hora_fin.strftime("%H:%M")}',
            reserva.get_estado_display(),
        ])
    if len(reservas_rows) == 1:
        reservas_rows.append(['-', 'Sin reservas', '-', '-', '-', '-'])
    reservas_table = build_table(reservas_rows, [2 * cm, 5.1 * cm, 5.1 * cm, 3 * cm, 4.1 * cm, 3.1 * cm], align='CENTER')
    reservas_table.setStyle(TableStyle([
        ('TEXTCOLOR', (-1, 1), (-1, -1), brand_dark),
    ]))
    elements.append(reservas_table)

    doc.build(elements, onFirstPage=draw_page_chrome, onLaterPages=draw_page_chrome)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_polideportivo_{timezone.localdate().strftime("%Y%m%d")}.pdf"'
    return response


@rol_requerido('administrador', 'municipio')
def exportar_excel_view(request):
    """Exporta un reporte Excel analitico sin seccion de puntos."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = _obtener_datos_reportes()
    wb = Workbook()

    title_font = Font(bold=True, size=15, color='183B56')
    subtitle_font = Font(size=10, color='64748B', italic=True)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='183B56', end_color='183B56', fill_type='solid')
    accent_fill = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')
    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    def configurar_hoja(ws, title, subtitle, total_columns):
        end_letter = get_column_letter(total_columns)
        ws.merge_cells(f'A1:{end_letter}1')
        ws.merge_cells(f'A2:{end_letter}2')
        ws['A1'] = title
        ws['A2'] = subtitle
        ws['A1'].font = title_font
        ws['A2'].font = subtitle_font
        ws['A1'].fill = accent_fill
        ws['A2'].fill = accent_fill
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.sheet_view.showGridLines = False

    def estilizar_encabezado(ws, row_index):
        for cell in ws[row_index]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def estilizar_cuerpo(ws, start_row, total_columns):
        for row_idx in range(start_row, ws.max_row + 1):
            for column_idx in range(1, total_columns + 1):
                cell = ws.cell(row=row_idx, column=column_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = even_fill

    def activar_filtros(ws, header_row, total_columns):
        end_letter = get_column_letter(total_columns)
        ws.freeze_panes = f'A{header_row + 1}'
        ws.auto_filter.ref = f'A{header_row}:{end_letter}{ws.max_row}'

    def ajustar_columnas(ws, total_columns):
        for column_idx in range(1, total_columns + 1):
            column_letter = get_column_letter(column_idx)
            max_length = 0
            for cell in ws[column_letter]:
                value = '' if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 38)

    ws1 = wb.active
    ws1.title = 'Resumen'
    configurar_hoja(ws1, 'Reporte del Polideportivo', f'Generado el {data["generated_at"].strftime("%d/%m/%Y %H:%M")}', 2)
    ws1.append([])
    ws1.append(['Metrica', 'Valor'])
    for label, value in data['resumen_metricas']:
        ws1.append([label, value])
    estilizar_encabezado(ws1, 4)
    estilizar_cuerpo(ws1, 5, 2)
    activar_filtros(ws1, 4, 2)
    ajustar_columnas(ws1, 2)

    ws2 = wb.create_sheet('Reservas')
    configurar_hoja(ws2, 'Reporte de reservas', 'Detalle completo de reservas registradas', 12)
    ws2.append([])
    ws2.append(['ID', 'Usuario', 'Correo', 'Cancha', 'Deporte', 'Fecha', 'Hora inicio', 'Hora fin', 'Estado', 'Invitados', 'Observacion', 'Creada'])
    for reserva in data['reservas_exportacion']:
        ws2.append([
            reserva.id,
            reserva.usuario.nombre,
            reserva.usuario.correo,
            reserva.cancha.nombre,
            reserva.cancha.tipo_deporte_label,
            reserva.fecha.strftime('%d/%m/%Y'),
            reserva.hora_inicio.strftime('%H:%M'),
            reserva.hora_fin.strftime('%H:%M'),
            reserva.get_estado_display(),
            reserva.lista_invitados or '',
            reserva.observacion or '',
            timezone.localtime(reserva.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
        ])
    estilizar_encabezado(ws2, 4)
    estilizar_cuerpo(ws2, 5, 12)
    activar_filtros(ws2, 4, 12)
    ajustar_columnas(ws2, 12)

    ws3 = wb.create_sheet('Usuarios')
    configurar_hoja(ws3, 'Reporte de usuarios', 'Usuarios registrados y estado actual', 8)
    ws3.append([])
    ws3.append(['ID', 'Nombre', 'Correo', 'Telefono', 'Rol', 'Estado', 'Fecha registro'])
    for usuario in data['usuarios_exportacion']:
        ws3.append([
            usuario.id,
            usuario.nombre,
            usuario.correo,
            usuario.telefono or '',
            usuario.get_rol_display(),
            usuario.puntos_acumulados,
            usuario.get_estado_display(),
            timezone.localtime(usuario.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
        ])
    estilizar_encabezado(ws3, 4)
    estilizar_cuerpo(ws3, 5, 7)
    activar_filtros(ws3, 4, 7)
    ajustar_columnas(ws3, 7)

    ws4 = wb.create_sheet('Canchas')
    configurar_hoja(ws4, 'Reporte de canchas', 'Inventario de canchas y configuracion deportiva', 7)
    ws4.append([])
    ws4.append(['ID', 'Nombre', 'Deporte', 'Ubicacion', 'Estado', 'Capacidad', 'Descripcion'])
    for cancha in data['canchas_exportacion']:
        ws4.append([
            cancha.id,
            cancha.nombre,
            cancha.tipo_deporte_label,
            cancha.ubicacion,
            cancha.get_estado_display(),
            cancha.capacidad_jugadores_label,
            cancha.descripcion or '',
        ])
    estilizar_encabezado(ws4, 4)
    estilizar_cuerpo(ws4, 5, 7)
    activar_filtros(ws4, 4, 7)
    ajustar_columnas(ws4, 7)

    ws5 = wb.create_sheet('Tendencias')
    configurar_hoja(ws5, 'Tendencias y analisis', 'Distribucion de reservas por estado, deporte y fechas recientes', 4)
    ws5.append([])
    ws5.append(['Categoria', 'Detalle', 'Valor', 'Participacion'])
    for item in data['estado_breakdown']:
        ws5.append(['Estado', item['label'], item['total'], f"{item['porcentaje']}%"])
    for item in data['deporte_breakdown']:
        ws5.append(['Deporte', item['label'], item['total'], f"{item['porcentaje']}%"])
    for item in data['top_canchas']:
        ws5.append(['Cancha', item['nombre'], item['total'], f"{item['porcentaje']}%"])
    for label, value in zip(data['chart_dias_labels'], data['chart_dias_data']):
        ws5.append(['Dia', label, value, ''])
    estilizar_encabezado(ws5, 4)
    estilizar_cuerpo(ws5, 5, 4)
    activar_filtros(ws5, 4, 4)
    ajustar_columnas(ws5, 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_polideportivo_{timezone.localdate().strftime("%Y%m%d")}.xlsx"'
    return response
